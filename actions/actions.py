# This files contains your custom actions which can be used to run
# custom Python code.
#
# See this guide on how to implement these action:
# https://rasa.com/docs/rasa/custom-actions


# This is a simple example for a custom action which utters "Hello World!"
import os.path
from typing import Any, Text, Dict, List
from openpyxl.reader.excel import ExcelReader
from pandas.io.pytables import AppendableFrameTable
from rasa_sdk import Action, Tracker
from rasa_sdk.executor import CollectingDispatcher
from rasa_sdk.events import SlotSet,AllSlotsReset
#from policies.policy import TestPolicy 
import pandas as pd
from pandas import ExcelWriter
import openpyxl as op
from openpyxl import load_workbook
from collections import OrderedDict
from GoogleDrive import *

from pathlib import Path
import os

#
#
# class ActionHelloWorld(Action):
#
#     def name(self) -> Text:
#         return "action_hello_world"
#
#     def run(self, dispatcher: CollectingDispatcher,
#             tracker: Tracker,
#             domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
#
#         dispatcher.utter_message(text="Hello World!")
#
#         return []
  
    
class imprimirSlot(Action):


    def get_project_root(self) -> Path:
        return Path(__file__).parent.parent

    def name(self) -> Text:
        return "action_imprimir_determinantes"
        
    def run(self, dispatcher: CollectingDispatcher,
        tracker: Tracker,
        domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        
        ######CONTENIDOS######:
        
        #1 figura humana completa
        _H = {"persona","humano","hombre","mujer","niño","niña","chico","chica","señor","señora","anciano","anciana","viejo","vieja","joven","nene","bebé","bebe","hombrecito","mujercita","personita","personas","humanos","hombres","mujeres","niños","niñas","chicos","chicas","señores","señoras","ancianos","ancianas","viejos","viejas","jovenes","nenes","bebés","bebes","hombrecitos","mujercitas","personitas","futbolista","jugador","jugadora","deportista","ladron","ladrón","ladrona","policía","futbolistas","deportistas","ladrones","ladronas","policias","jugadores","jugadoras","guerrero","guerreros","guerrera","guerreras"}
        #2 figura humana completa irreal, de ficción o mitológica
        _ParentesisH = {"payaso","payasos","hada","hadas","bruja","brujas","fantasma","fantasmas","enano","enanos","enana","enanas","demonio","demonios","ángel","ángeles","humanoide","humanoides","caricaturas","caricatura","monstruo","monstruos","monstruito","monstruitos","duende","duendes"}
        #3 Detalle humano
        _Hd = {"brazo","pierna","dedos","pies","cabeza","codo","nariz","brazos","piernas","diente","dientes","muela","muelas","rodilla","rodillas","cerebro","cerebros",}        
        #4 Detalle humano irreal, de ficción o mitológico
        _ParentesisHd = {"máscara","mascara","máscaras","mascaras"}
        #5 Experiencia humana
        _Hx = {"amor","amar","ama","amamos","amo","odio","odia","odiamos","odiar","depresión","deprimido","deprimida","deprimidos","deprimidas","feliz","felices","alegre","alegres","felicidad","alegria","ruido","ruidoso","sonido","suena","olor","huele","oloroso","miedo","temor","miedoso","contento","contenta","contentos","contentas"}
        #6 Figura animal completa
        _A = {"escarabajo","escarabajos","bicho","bichos","araña","arañas","cucaracha","cucarachas","mariposa","mariposas","mantis","mosca","mosquito","moscas","mosquitos","murcielago","murciélago","murciélagos","pulga","pulgas","águila","águilas","avestruz","ballena","bisonte","bug","búfalo","búhos","buitre","burro","caballo","cabra","camaleón","camello","canario","castor","cebra","cerdo","chancho","ciervo","cobra","colibrí","comadreja","cóndor","conejo","delfín","elefante","faisan","flamenco","foca","gallina","gallo","gato","gorila","guepardo","hámster","hiena","hipopótamo","jabalí","jaguar","jirafa","koala","lagarto","león","lobo","loro","manatí","mapache","mono","murciélago","nutria","ñandú","orcas","oso","pájaro","paloma","panda","pato","pavo","pelícano","perro","pingüino","puercoespín","puma","rana","ratón","reno","rinoceronte","salamandra","sapo","serpiente","tapir","tejon","tiburón","tigre","topo","toro","tucán","vaca","vicuña","zorrino","zorro","águila","avestruces","ballenas","bisontes","búfalos","bugs","búho","buitres","burros","caballos","cabras","camaleones","camellos","canarios","castores","cebras","cerdos","chanchos","ciervos","cobras","colibries","comadrejas","cóndores","conejos","delfines","elefantes","faisanes","flamencos","focas","gallinas","gallos","gatos","gorilas","guepardos","hámsters","hienas","hipopótamos","jabalies","jaguares","jirafas","koalas","lagartos","leones","lobos","loros","manaties","mapaches","monos","murciélagos","nutrias","ñandues","orca","osos","pájaros","palomas","pandas","patos","pavos","pelícanos","perros","pingüinos","puercoespines","pumas","ranas","ratones","renos","rinocerontes","salamandras","sapos","serpientes","tapires","tiburones","tigres","topos","toros","tucanes","vacas","víbora","víboras","vicuñas","zorrinos","zorros""bacteria","bacterias","animal","animales","arácnido","aracnido","arácnidos","arácnido","libélula","libélulas","libelula","libelulas","ciempies","pescado","bagre","escorpión","escorpion","escorpiones","cotorras","caballito","elefantito","camaron","camarones","artrópodo","artropodo","artrópodos","artropodos","garrapata","garrapatas","peces","pez","langosta","langostas"}
        #7 Figura animal completa irreal, de ficción o mitológica
        _ParentesisA = {"fiera","bruja","brujas","ángel","angel","ángeles","angeles","demonio","demonios","fantasma","fantasmas","unicornio","unicornios","dragón","dragon","dragones","minotauro","minotauros","krampus","hada","guason","batman","superman","korioto","koriotos","pato donald","mickey","yeti","yetis",}
        #8 Figura animal incompleta
        _Ad = {"pata","patas","cola","pinza","hocico", "cuero","pezuñas","garras","vasos","pico","melena","trompa","cuerno","colmillo","colmillos"}
        #9 Figura animal irreal, de ficción o mitológica incompleta. PENDIENTE
        #10 Anatomía
        _An = {"ósea","osea","cráneo","cráneo","torax","toracica","tórax","corazón","corazon","pulmón","pulmon","estomágo","estomago","panza","hígado","higado","musculo","articulaciones","vértebral","vértebra","vertebras","cerebro","cerebros"}
        #11 Arte
        _Art = {"pintura","dibujo","pinturas","dibujos","ilustración","ilustracion","ilustraciones","acuarela","acuarelas","arte","estatua","estatuas","escultura","esculturas","joya","joyas","insignia","insignias","escudo","escudos","adornos","espada","espadas","cuadros","lienzos","cuadro","lienzo","logo","logos"}
        #12 Antropología
        _Ay = {"tótem","totem","templo","prehistórica","prehistorica","prehistóricas","prehistorica","incaica","incaicas","azteca","aztecas","maya","mayas","romano","romanos","griego","griegos","persa","persas"}
        #13 Sangre
        _Bl = {"sangre","sanguíneo","sanguínea","sanguinario","sanguinaria","sangriento","sangrienta"}
        #14 Botánica
        _Bt = {"vegetal","vegetales","arbusto","arbustos","flor","flores","floral","alga","algas","arbol","árbol","árboles","arboles","hoja","hojas","pétalo","pétalos","tronco","tronco","raiz","raíces","nido","hongo","hongos","pino","palmera","pinos","palmeras"} 
        #15 Vestidos
        _Cg = {"sombrero","sombreros","gorro","gorros","gorras","gorra","bota","botas","botines","cinturón","cinturon","cinturones","corbata","corbatas","moño","moños","chaqueta","chaquetas","saco","sacos","polera","poleras","pantalón","pantalon","pantalones","bufanda","bufandas","bermudas","short","shorts","medias","calzoncillos","musculosa"}
        #16 Nubes
        _Cl = {"nubes","nube","nubarrón","nubarrones","nublado","nublada"}
        #17 Explosión
        _Ex = {"explosión","explosión","explotar","implotar","bomba","explosivos","estallido","estallidos","detonación","detonaciones","bombardeo","bombardeos"}
        #18 Fuego
        _Fi = {"fuego","fuegos","llama","llamas","incendio","incendios","humo","humos","fogata","fogatas","chispa","chispas","hoguera","hogueras","quema","quemas"}
        #19 Comida
        _Fd = {"pollo","pollos","carne","carnes","pescado","pescados","helado","helados","panqueque","panqueques","verdura","verduras","papa","papas","zapallo","zapallos","lechuga","lechugas","tomates","tomate","zanahoria","zanahorias","sandía","sandías","melón","melones","kiwi","kiwis","mandarina","naranja","mandarinas","naranjas","banana","bananas","palta","frutas","verdura","fruta","torta","budín","tostadas","sandwich","sandwiches","pan","panes","sopa","pasta","sopas","pastas","spaghetti"}
        #20 Geografía
        _Ge = {"mapa","mapas","plano","planos","cartográfico","cartografía","continente","continentes","país","paises","ciudad","ciudad","region","regiones"}
        #21 Hogar
        _Hh = {"cama","cucheta","sommier","sillón","silla","mesa","lámpara","cuchillo","olla","alfombra","cortina","mueble","horno","cocina","pieza","habitación","baño","cochera","garage","camas","cuchetas","sommiers","sillones","sillas","mesas","lámparas","cuchillos","ollas","alfombras","cortinas","muebles","hornos","cocinas","piezas","habitaciones","baños","cocheras"}
        #22 Paisaje
        _Ls = {"montaña","montañas","cordillera","cordilleras","colina","colinas","cerro","cerros","sierra","sierras","isla","islas","cueva","cuevas","roca","rocas","piedra","piedras","bosque","bosques","desierto","desierto","llanura","llanuras","pantano","pantanos","glaciar","glaciares"}
        #23 Naturaleza
        _Na = {"lago","laguna","lagos","lagunas","sol","luna","mar","mar","océano","océanos","agua","hielo","hielos","lluvia","lluvias","niebla","neblina","bruma","tormenta","brisa","viento","arco iris","tormenta","noche","día","dia","granizo","trueno","relámpago","relampago","bosque"}
        #24 Ciencia
        _Sc = {"avión","aviones","avion","edificio","edificios","puente","puentes","auto","autos","moto","motos","microscopio","microscopios","laboratorio","laboratorios","motor","motores","turbina","turbinas","telescopio","telescopios","arma","armas","armamento","cohete","cohetes","nave","naves","ovni","OVNI","ovnis","OVNIS","barco","barcos","antena","antenas","satélite","satélites","satelite","satelites"}
        #25 Sexo
        _Sx = {"pene","penes","verga","vergas","pito","pitos","vagina","vaginas","concha","conchas","nalgas","cachas","pechos","teta","tetas","testículos","huevos","bolas","menstruación","aborto","abortar","coito","coger","garchar","cogiendo","teniendo sexo","garchando"}
        #26 Radiografía
        _Xy = {"radiografía","radiografia","placa","placas","rayos x","tomografía","ecografía","tomografía","ultrasonido","resonancia"}
        #27 Popular1
        _Po1 = {"escarabajo","escarabajos","bicho","bichos","araña","arañas","cucaracha","cucarachas","mariposa","mariposas","mantis","mosca","mosquito","moscas","mosquitos","insecto","insectos","gusano"}
        #28 Popular3
        _Po3 = {"persona","humano","hombre","mujer","niño","niña","chico","chica","señor","señora","personas","humanos","hombres","mujeres","niños","niñas","chicos","chicas","señores","señoras","payaso","payasos","hada","hadas","bruja","brujas","fantasma","fantasmas","enano","enanos","enana","enanas","demonio","demonios","ángel","ángeles","humanoide","humanoides","caricaturas","caricatura","monstruo","monstruos","duende","duendes"}
        #IMAGEN 2 NO TIENE RESPUESTAS POPULARES
        
        slot_nombre = str(tracker.get_slot("nombre")).replace(' ','')
        SlotSet("nombre",slot_nombre)
        #tomo planilla de excel
        wb = op.load_workbook(str(self.get_project_root())+os.path.sep +'files'+os.path.sep+slot_nombre+'.xlsx')
        ws = wb.get_sheet_by_name('Hoja de datos')
        determinantes = '?'        
        popular = '?'
        contenidos = ''
        lamina = tracker.get_slot("contador")
        if lamina == 4:
            contenidos = ws['I'+str(lamina - 2)].value
            lamina = 1
        elif lamina == 5:
            contenidos = ws['I'+str(lamina - 2)].value
            lamina = 2
        elif lamina == 6:
            contenidos = ws['I'+str(lamina - 2)].value
            lamina = 3
        respuesta = tracker.latest_message['text']
        contenidos = str(contenidos)
        while _H:#1 figura humana completa
            subconjunto = _H.pop()
            if subconjunto in respuesta:
                if ' H,' not in contenidos:
                    contenidos = contenidos + ' H,'
                elif contenidos == '':
                    contenidos = ' H,'    
                break   
        while _ParentesisH:#2 figura humana completa irreal, de ficción o mitológica
            subconjunto = _ParentesisH.pop()
            if subconjunto in respuesta:
                if ' (H),' not in contenidos:
                    contenidos = contenidos + ' (H),'
                elif contenidos == '':
                    contenidos = ' (H),'    
                break
        while _Hd: #3 Detalle humano
            subconjunto = _Hd.pop()
            if subconjunto in respuesta:
                if ' Hd,' not in contenidos:
                    contenidos = contenidos + ' Hd,'
                elif contenidos == '':
                    contenidos = ' Hd,'    
                break
        while _ParentesisHd: #4 Detalle humano irreal, de ficción o mitológico
            subconjunto = _ParentesisHd.pop()
            if subconjunto in respuesta:
                if ' (Hd),' not in contenidos:
                    contenidos = contenidos + ' (Hd),'
                elif contenidos == '':
                    contenidos = ' (Hd),'    
                break    
        while _Hx: #5 Experiencia humana
            subconjunto = _Hx.pop()
            if subconjunto in respuesta:
                if ' Hx,' not in contenidos:
                    contenidos = contenidos + ' Hx,'
                elif contenidos == '':
                    contenidos = ' Hx,'    
                break
        while _A: #6 Figura animal completa
            subconjunto = _A.pop()
            if subconjunto in respuesta:
                if ' A,' not in contenidos:
                    contenidos = contenidos + ' A,'
                elif contenidos == '':
                    contenidos = ' A,'    
                break
        while _ParentesisA: #7 Figura animal completa irreal, de ficción o mitológica
            subconjunto = _ParentesisA.pop()
            if subconjunto in respuesta:
                if ' (A),' not in contenidos:
                    contenidos = contenidos + ' (A),'
                elif contenidos == '':
                    contenidos = ' (A),'    
                break
        while _Ad: #8 Figura animal incompleta
            subconjunto = _Ad.pop()
            if subconjunto in respuesta:
                if ' Ad,' not in contenidos:
                    contenidos = contenidos + ' Ad,'
                elif contenidos == '':
                    contenidos = ' Ad,'    
                break
        while _An: #10 Anatomía
            subconjunto = _An.pop()
            if subconjunto in respuesta:
                if ' An,' not in contenidos:
                    contenidos = contenidos + ' An,'
                elif contenidos == '':
                    contenidos = ' An,'    
                break
        while _Art: #11 Arte
            subconjunto = _Art.pop()
            if subconjunto in respuesta:
                if ' Art,' not in contenidos:
                    contenidos = contenidos + ' Art,'
                elif contenidos == '':
                    contenidos = ' Art,'    
                break
        while _Ay: #12 Antropológica
            subconjunto = _Ay.pop()
            if subconjunto in respuesta:
                if ' Ay,' not in contenidos:
                    contenidos = contenidos + ' Ay,'
                elif contenidos == '':
                    contenidos = ' Ay,'    
                break
        while _Bl: #13 Sangre
            subconjunto = _Bl.pop()
            if subconjunto in respuesta:
                if ' Bl,' not in contenidos:
                    contenidos = contenidos + ' Bl,'
                elif contenidos == '':
                    contenidos = ' Bl,'    
                break
        while _Bt: #14 Botánica
            subconjunto = _Bt.pop()
            if subconjunto in respuesta:
                if ' Bt,' not in contenidos:
                    contenidos = contenidos + ' Bt,'
                elif contenidos == '':
                    contenidos = ' Bt,'    
                break
        while _Cg: #15 Vestidos
            subconjunto = _Cg.pop()
            if subconjunto in respuesta:
                if ' Cg,' not in contenidos:
                    contenidos = contenidos + ' Cg,'
                elif contenidos == '':
                    contenidos = ' Cg,'    
                break
        while _Cl: #16 Nubes
            subconjunto = _Cl.pop()
            if subconjunto in respuesta:
                if ' Cl,' not in contenidos:
                    contenidos = contenidos + ' Cl,'
                elif contenidos == '':
                    contenidos = ' Cl,'    
                break
        while _Ex: #17 Explosión
            subconjunto = _Ex.pop()
            if subconjunto in respuesta:
                if ' Ex,' not in contenidos:
                    contenidos = contenidos + ' Ex,'
                elif contenidos == '':
                    contenidos = ' Ex,'    
                break
        while _Fi: #18 Fuego
            subconjunto = _Fi.pop()
            if subconjunto in respuesta:
                if ' Fi,' not in contenidos:
                    contenidos = contenidos + ' Fi,'
                elif contenidos == '':
                    contenidos = ' Fi,'    
                break
        while _Fd: #19 Comida
            subconjunto = _Fd.pop()
            if subconjunto in respuesta:
                if ' Fd,' not in contenidos:
                    contenidos = contenidos + ' Fd,'
                elif contenidos == '':
                    contenidos = ' Fd,'    
                break
        while _Ge: #20 Geografía
            subconjunto = _Ge.pop()
            if subconjunto in respuesta:
                if ' Ge,' not in contenidos:
                    contenidos = contenidos + ' Ge,'
                elif contenidos == '':
                    contenidos = ' Ge,'    
                break
        while _Hh: #21 Hogar
            subconjunto = _Hh.pop()
            if subconjunto in respuesta:
                if ' Hh,' not in contenidos:
                    contenidos = contenidos + ' Hh,'
                elif contenidos == '':
                    contenidos = ' Hh,'    
                break
        while _Ls: #21 Paisaje
            subconjunto = _Ls.pop()
            if subconjunto in respuesta:
                if ' Ls,' not in contenidos:
                    contenidos = contenidos + ' Ls,'
                elif contenidos == '':
                    contenidos = ' Ls,'    
                break
        while _Na: #23 Naturaleza
            subconjunto = _Na.pop()
            if subconjunto in respuesta:
                if ' Na,' not in contenidos:
                    contenidos = contenidos + ' Na,'
                elif contenidos == '':
                    contenidos = ' Na,'    
                break
        while _Sc: #24 Ciencia
            subconjunto = _Sc.pop()
            if subconjunto in respuesta:
                if ' Sc,' not in contenidos:
                    contenidos = contenidos + ' Sc,'
                elif contenidos == '':
                    contenidos = ' Sc,'    
                break
        while _Sx: #25 sexo
            subconjunto = _Sx.pop()
            if subconjunto in respuesta:
                if ' Sx,' not in contenidos:
                    contenidos = contenidos + ' Sx,'
                elif contenidos == '':
                    contenidos = ' Sx,'    
                break
        while _Xy: #26 Radiografía
            subconjunto = _Xy.pop()
            if subconjunto in respuesta:
                if ' Xy,' not in contenidos:
                    contenidos = contenidos + ' Xy,'
                elif contenidos == '':
                    contenidos = ' Xy,'    
                break
        lamina = tracker.get_slot("contador")
        while _Po1: #27 populares
            subconjunto = _Po1.pop()
            if subconjunto in respuesta:
                if lamina == 1 or lamina == 4:#and zona corresponde
                    popular = 'Po1' #hay que revisar si coincide con sector de la imagen
                    break
                else: 
                    popular = '?'            

        #IMAGEN 2 NO TIENE RESPUESTAS POPULARES

        while _Po3: #28 populares
            subconjunto = _Po3.pop()
            if subconjunto in respuesta:
                if lamina == 3 or lamina == 6: #and zona corresponde
                    popular = 'Po3'       
                    break
                else: popular ='?'
        #####comprobar populares al final según orden de las respuestas

        ##todo lo que no entra en un conjunto es contenido ideográfico "Id"
        sender = tracker.sender_id


        ######DETERMINANTES#####
        slot_paridad = tracker.get_slot("par")
        slot_vista = tracker.get_slot("vista")
        slot_ccromatico = tracker.get_slot("color_cromatico")
        slot_cacromatico = tracker.get_slot("color_acromatico")
        slot_forma = tracker.get_slot("forma")
        slot_movimiento = tracker.get_slot("movimiento")
        slot_textura = tracker.get_slot("textura")
        slot_reflejo = tracker.get_slot("reflejo") 
        slot_sombreado = tracker.get_slot("sombreado")      
        
        if slot_paridad == "true":
            par = '2'
        else:
            par ='?'
        if lamina == 4:
            determinantes = ws['F'+str(lamina - 2)].value
            lamina = 1
        elif lamina == 5:
            determinantes = ws['F'+str(lamina - 2)].value
            lamina = 2
        elif lamina == 6:
            determinantes = ws['F'+str(lamina - 2)].value
            lamina = 3


        if slot_vista == "true":
            if determinantes == '?':
                determinantes = ' V,'
            elif ' V,' not in determinantes:
                determinantes = determinantes + ', V,'
        if slot_ccromatico == "true":
            if determinantes == '?':
                determinantes = ' C,'
            elif ' C,' not in determinantes:
                determinantes = determinantes + ' C,'
        if slot_cacromatico == "true":
            if determinantes == '?':
                determinantes = ' C\','
            elif ' C\',' not in determinantes:
                determinantes = determinantes + ' C\','
        if ((slot_forma == "humana") or (slot_forma == "animal") or (slot_forma == "inanimada")):
            if determinantes == '?':
                determinantes = ' F,'
            elif ' F,' not in determinantes and' F,' not in ws['F'+ str(lamina)].value: 
                determinantes = determinantes + ' F,'
        if slot_movimiento == "true": 
            if slot_forma == 'humana':
                if determinantes == '?':
                    determinantes = ' M,'
                elif ' M,' not in determinantes:
                    determinantes = determinantes + ' M,'
            elif slot_forma == 'inanimada':
                if determinantes == '?':
                    determinantes = ' m,'
                elif ' m,' not in determinantes: 
                    determinantes = determinantes + ' m,'
            else:
                if determinantes == '?':
                    determinantes = ' ind,'
                elif ' ind,' not in determinantes:
                    determinantes = determinantes + ' ind,'
        if slot_textura == "true":
            if determinantes == '?':
                determinantes = ' T,'
            elif ' T,' not in determinantes:
                determinantes = determinantes + ' T,'
        if slot_reflejo == "true":
            if determinantes == '?':
                determinantes = ' r,'
            elif ' r,' not in determinantes:
                determinantes = determinantes + ' r,'
        if slot_sombreado == "true":
            if determinantes == '?':
                determinantes = ' Y,'
            elif ' Y,' not in determinantes:
                determinantes = determinantes + ' Y,'
        
        lamina = tracker.get_slot("contador")
        # Lo que hace es mostrar el mensaje con la próxima imagen: utter_Lamina2 ó utter_Lamina3
        next_response = tracker.get_slot("response")
        if next_response != "None":
            dispatcher.utter_message(response=next_response)  

        
        if (lamina == 1) or (lamina == 2) or (lamina == 3):
            indice = str(lamina + 1)
            ws['A' + indice]=lamina
            ws['B' + indice]='1'
            ws['C' + indice]='?'
            ws['D' + indice]='?'
            ws['E' + indice]='?'
            ws['F' + indice]=determinantes
            ws['G' + indice]='?'
            ws['H' + indice]=par
            ws['I' + indice]=contenidos
            ws['J' + indice]=popular
            ws['K' + indice]='?'
            ws['L' + indice]='?'
            ws['M' + indice]=respuesta
            ws['N' + indice]='?'
            #ws.append([lamina,'1','?','?','?', determinantes,'?', par, contenidos, popular,'?','?']) 
        else:
            if lamina == 4:
                ws['N2']=respuesta
                if determinantes != '?':
                    ws['F2'] = (determinantes)[:-1]
                if  ws['H2'] == '2':
                    ws['H2'] = par
                if contenidos != '?':  
                    ws['I2'] = contenidos[:-1]
                vj2 = ws['J2'].value
                if str(vj2) !='?' and str(vj2) != 'Po1':
                    ws ['J2'] = str(vj2) + popular
            elif lamina == 5:
                ws['N3']=respuesta
                if determinantes != '?':
                    ws['F3'] = (determinantes)[:-1]
                if  ws['H3'] == '2':
                    ws['H3'] = par
                if contenidos != '?':
                    ws['I3'] =  contenidos[:-1]
            elif lamina == 6:
                ws['N4']=respuesta
                if determinantes != '?':
                    ws['F4'] = (determinantes)[:-1]
                if ws['H4'] == '?':
                    ws['H4'] = par
                ws['I4'] = contenidos[:-1]
                vj4 = ws['J4'].value
                if str(vj4) !='?' and str(vj4) != 'Po3':
                    ws['J4'] = str(vj4) + popular
        wb.save(str(self.get_project_root())+os.path.sep +'files'+os.path.sep+slot_nombre+'.xlsx')
        wb.close()
        if lamina == 6:
            
            subir_archivo(str(self.get_project_root())+os.path.sep +'files'+os.path.sep+slot_nombre+'.xlsx',"1EQ4h-Blfc3PqySXRvViSVrq2ZhCmq2rl")
        planilla = pd.read_excel(str(self.get_project_root())+os.path.sep +'files'+os.path.sep+slot_nombre+'.xlsx')
        print(planilla)
        return [SlotSet("par","false"),SlotSet("vista","false"),SlotSet("color_cromatico","false"),SlotSet("color_acromatico","false"),SlotSet("forma","false"),SlotSet("movimiento","false"),SlotSet("textura","false"),SlotSet("reflejo","false"),SlotSet("sombreado","false"),SlotSet("response", "None")]

class TercerParte(Action):

     def name(self) -> Text:
         return "action_TerceraParte"

     def run(self, dispatcher: CollectingDispatcher,
             tracker: Tracker,
             domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:

        next_response = tracker.get_slot("response")
        dispatcher.utter_message(response=next_response)  
        return []